#!/usr/bin/env python

"""
Python script to collect the output of the SHOW VPN-SESSIONDB ANYCONNECT command from a
set of Cisco ASA firewalls. The output is then saved to a Microsoft Excel file.
"""

from datetime import datetime
from getpass import getpass
from netmiko import ConnectHandler
from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import get_column_letter
from openpyxl.worksheet.table import Table


def main():
    """
    The main() function has the following duties:
     - Contains the inventory of Cisco ASA firewalls to be queried.
     - Obtains the current date and time.
     - Calls a function to retrieve credentials for the firewalls.
     - Iterates through the inventory, calling a function to retieve the desired
       information.
     - Calls a function to save the information to a Microsoft Excel file.
    """

    username, password = get_creds()

    primary = {
        "device_type": "cisco_asa",
        "host": "fwl-dc1-vpn-a",
        "username": username,
        "password": password,
    }

    secondary = {
        "device_type": "cisco_asa",
        "host": "fwl-dc0-inet-a",
        "username": username,
        "password": password,
    }

    now = datetime.now()
    tab_name = now.strftime("%Y_%m_%d_%H_%M_%S")

    results = []
    for firewall in [primary, secondary]:
        results.append(firewall["host"])
        vpn_sessiondb = show_vpn_sessiondb(firewall)
        results.append(vpn_sessiondb)

    output_to_excel(tab_name, results)


def get_creds():
    """
    The get_creds() funtion queries the user for the username and password needed to
    logon to each firewall.
    """

    print("-" * 40)

    un = input("Username, (q) to quit: ")
    if un.lower() == "q":
        exit("QUITTING")
    pw = getpass()
    if pw.lower() == "q":
        exit("QUITTING")

    print("-" * 40)

    return (un, pw)


def show_vpn_sessiondb(device):
    """
    The show_vpn_sessiondb() function uses Netmiko to connect to each firewall, and
    collect the output from the SHOW VPN-SESSIONDB ANYCONNECT command. It uses TextFSM
    from network.toCode() to convert the output from one large string to structured data.

    ARGS:
        device (Dictionary): Device information used by Netmiko.
    """

    done = False
    while not done:
        try:
            net_connect = ConnectHandler(**device)
            print(f"Gathering information from {device['host']}")
            done = True
        except:
            print("\n")
            print(f"ERROR: Invalid username or password for {device['host']}")
            username, password = get_creds()
            device["username"] = username
            device["password"] = password

    output = net_connect.send_command("show vpn-sessiondb anyconnect", use_textfsm=True)
    net_connect.disconnect()
    return output


def output_to_excel(tab, data):
    """
    The output_to_excel() function saves information to a Microsoft Excel file.

    ARGS:
        tab (String): Current date and time used to create the tab name and the table
        name in the spreadsheet.

        data (List): The data saved to a spreadsheet.
    """

    PATH = r"S:\Cit\Operations\Network\AnyConnect"
    FILE = r"\AnyConnect.xlsx"
    excel_file = PATH + FILE

    try:
        wb = load_workbook(filename=excel_file)
    except:
        wb = Workbook()
        ws = wb.active
        wb.remove(ws)

    ws = wb.create_sheet(tab, 0)

    ws.cell(1, 1, "Firewall")
    column_number = {}
    # Using the first dictionary from the second item
    # in the list to generate the column headings
    headings = data[1][0].keys()
    number_of_columns = len(headings) + 1
    max_column = get_column_letter(number_of_columns)
    for element, heading in enumerate(headings, 2):
        # Populating the column header cells
        ws.cell(1, element, heading)
        # Creating a dictionary to store the column headers
        # with their column number because dictionaries are unordered
        column_number[heading] = element

    row_number = 1
    # Iterating through the list
    for item in data:
        if type(item) == str:
            hostname = item
        else:
            # Iterating through the list of dictionaries
            for row_data in item:
                row_number += 1
                # Populating the first cell in the row with the firewall name
                ws.cell(row_number, 1, hostname)
                for column_heading in headings:
                    # Populating the rest of the cells in the row by using
                    # a dictionary lookup where the column header is the key
                    ws.cell(
                        row_number,
                        column_number.get(column_heading),
                        row_data.get(column_heading),
                    )

    table_ref = f"A1:{max_column}{row_number}"
    table_name = f"_{tab}"
    vpn_table = Table(displayName=table_name, ref=table_ref)
    ws.add_table(vpn_table)
    ws.freeze_panes = "D2"
    wb.save(excel_file)
    print(f"Recorded {row_number} rows in spreadsheet {excel_file}")


if __name__ == "__main__":
    main()
